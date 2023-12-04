import boto3
import openpyxl

# Function to update or append data in the sheet
def update_or_append_data(sheet, account_id, domain_name, number_of_nodes, instance_type):
    row_idx = 2  # Start from the second row, skipping the header
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        if row[0] == account_id and row[1] == domain_name:
            # Update existing row if found
            print(f"Found existing row with domain {domain_name} in account {account_id}. Updating the same row.")
            sheet.cell(row=row_idx, column=3, value=number_of_nodes)
            sheet.cell(row=row_idx, column=4, value=instance_type)
            return
        row_idx += 1

    # Append new row if not found
    print(f"Adding row {[account_id, domain_name, number_of_nodes, instance_type]}")
    sheet.append([account_id, domain_name, number_of_nodes, instance_type])

# Initialize Boto3 clients
client = boto3.client('opensearch')
sts_client = boto3.client('sts')

# Get the current AWS account ID
account_id = sts_client.get_caller_identity().get('Account')

# Create or load an Excel workbook and sheet
try:
    wb = openpyxl.load_workbook('OpenSearchDomains.xlsx')
    sheet = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['Account ID', 'Domain Name', 'Number of Nodes', 'Instance Type','Costs'])

# List all OpenSearch domains
response = client.list_domain_names()

for domain_info in response['DomainNames']:
    domain_name = domain_info['DomainName']

    # Get detailed information about each domain
    domain_config = client.describe_domain(DomainName=domain_name)
    cluster_config = domain_config['DomainStatus']['ClusterConfig']

    number_of_nodes = cluster_config['InstanceCount']
    instance_type = cluster_config['InstanceType']

    # Update or append the domain details to the spreadsheet
    update_or_append_data(sheet, account_id, domain_name, number_of_nodes, instance_type)

# Save the Excel workbook
wb.save('OpenSearchDomains.xlsx')
print("Workbook saved.")